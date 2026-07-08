from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


class ExcelService:
    """
    Generic Excel Helper
    Digunakan oleh seluruh export report.
    """

    HEADER_FILL = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    HEADER_FONT = Font(
        bold=True,
        color="FFFFFF"
    )

    BODY_FONT = Font(
        bold=False,
        color="000000"
    )

    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    CENTER = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    LEFT = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )

    @staticmethod
    def create_workbook(title: str):
        """
        Create workbook & active worksheet.
        """
        wb = Workbook()

        ws = wb.active
        ws.title = title

        return wb, ws

    @staticmethod
    def write_headers(ws, headers: list):
        """
        Write header row.
        """

        for col, header in enumerate(headers, start=1):

            cell = ws.cell(row=1, column=col)

            cell.value = header

            cell.font = ExcelService.HEADER_FONT
            cell.fill = ExcelService.HEADER_FILL
            cell.border = ExcelService.THIN_BORDER
            cell.alignment = ExcelService.CENTER

    @staticmethod
    def write_rows(ws, rows: list):
        """
        Write table body.
        """

        for row_idx, row in enumerate(rows, start=2):

            for col_idx, value in enumerate(row, start=1):

                cell = ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value=value,
                )

                cell.font = ExcelService.BODY_FONT
                cell.border = ExcelService.THIN_BORDER
                cell.alignment = ExcelService.LEFT

    @staticmethod
    def auto_width(ws):
        """
        Auto fit column width.
        """

        for column_cells in ws.columns:

            max_length = 0

            column = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:

                try:

                    if cell.value:

                        max_length = max(
                            max_length,
                            len(str(cell.value)),
                        )

                except Exception:
                    pass

            ws.column_dimensions[column].width = max_length + 3

    @staticmethod
    def to_bytes(workbook):
        """
        Convert workbook to BytesIO.
        """

        output = BytesIO()

        workbook.save(output)

        output.seek(0)

        return output